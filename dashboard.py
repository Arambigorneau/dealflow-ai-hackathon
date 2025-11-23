import time
import PyPDF2
import json
import os
import re
import pandas as pd
import altair as alt
import streamlit as st
import streamlit.components.v1 as components
from ibm_watsonx_ai.foundation_models import ModelInference
from ibm_watsonx_ai.metanames import GenTextParamsMetaNames as GenParams
from ibm_watsonx_ai import Credentials

# ==========================================
# 1. CONFIGURATION
# ==========================================
st.set_page_config(
    page_title="DealFlow AI | M&A Audit",
    page_icon="⚖️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Initialize session state
if "uploader_key" not in st.session_state: st.session_state.uploader_key = 0
if "audit_stats" not in st.session_state: st.session_state.audit_stats = None
if "audit_db" not in st.session_state: st.session_state.audit_db = {}
if "chat_history" not in st.session_state: st.session_state.chat_history = []
if "full_corpus" not in st.session_state: st.session_state.full_corpus = ""

# ==========================================
# 2. CSS & UI (merged / preserved)
# ==========================================
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
    
    html, body, [class*="css"] {
        font-family: 'Inter', sans-serif;
    }

    /* --- SIDEBAR COMPACTAGE & WIDGETS --- */
    section[data-testid="stSidebar"] .block-container {
        padding-top: 0.5rem !important;
        padding-bottom: 0.5rem !important;
        gap: 0.6rem !important;
    }
    section[data-testid="stSidebar"] div[data-testid="stVerticalBlock"] {
        gap: 0.4rem !important;
    }
    [data-testid="stFileUploader"] section[aria-label="File list"] {
        max-height: 280px !important;
        overflow-y: auto !important;
    }
    [data-testid="stFileUploader"] li {
        padding: 1px 0px !important;
        margin: 0px 0 !important;
    }

    /* Reduce spacing around sidebar buttons aggressively */
    section[data-testid="stSidebar"] div.stButton,
    section[data-testid="stSidebar"] div[data-testid="stDownloadButton"],
    section[data-testid="stSidebar"] div.row-widget,
    section[data-testid="stSidebar"] div.element-container {
        margin-top: 0 !important;
        margin-bottom: 0 !important;
        padding-top: 0 !important;
        padding-bottom: 0 !important;
        gap: 0 !important;
    }
    section[data-testid="stSidebar"] button {
        padding: 0.1rem 0.4rem !important;
        font-size: 0.85rem !important;
        margin: 0 !important;
    }

    /* Tab highlight and active tab color (IBM blue) */
    div[data-baseweb="tab-highlight"] {
        background-color: #0f62fe !important;
    }
    button[data-baseweb="tab"][aria-selected="true"] {
        color: #0f62fe !important;
    }
    button[data-baseweb="tab"]:hover { color: #0f62fe !important; }
    
    /* Radio buttons styling: outer ring blue when checked */
    div[role="radiogroup"] label input:checked ~ div > div:first-child {
        border-color: #0f62fe !important;
        background-color: transparent !important;
    }
    div[role="radiogroup"] label input:checked ~ div > div:first-child > div {
        background-color: #0f62fe !important;
        box-shadow: none !important;
    }
    /* Make label text of selected radio blue */
    div[role="radiogroup"] label input:checked ~ div > div:last-child,
    div[role="radiogroup"] label[aria-checked="true"] {
        color: #0f62fe !important;
        font-weight: 600;
    }

    /* Metrics / delta styling: make delta background transparent */
    [data-testid="stMetricDelta"], [data-testid="stMetricDelta"] > div {
        background-color: transparent !important;
        box-shadow: none !important;
    }
    [data-testid="stMetricDelta"] svg { fill: #0f62fe !important; }
    [data-testid="stMetricDelta"] p, [data-testid="stMetricDelta"] span { color: #0f62fe !important; }

    /* === Forcer le texte des deltas KPI (ex: "Requires Attention") en blanc === */
    [data-testid="stMetricDelta"],
    [data-testid="stMetricDelta"] > div,
    [data-testid="stMetricDelta"] span,
    [data-testid="stMetricDelta"] p,
    [data-testid="stMetricDelta"] svg {
        color: blue !important;
        fill: blue !important;
        background: transparent !important;
    }

    /* === Forcer le texte des filtres / radios en blanc (sélectionné ou non) === */
    div[role="radiogroup"] label,
    div[role="radiogroup"] label * ,
    div[role="radiogroup"] label > div:last-child,
    div[role="radiogroup"] label[aria-checked="true"],
    div[role="radiogroup"] label input ~ div {
        color: white !important;
    }

    /* Au cas où les filtres sont rendus comme boutons tabs, forcer aussi ces sélecteurs */
    button[data-baseweb="tab"] p,
    button[data-baseweb="tab"] {
        color: white !important;
    }

    /* Assurer que le ring extérieur (checked) reste visible mais texte blanc */
    div[role="radiogroup"] label input:checked ~ div > div:first-child {
        border-color: #0f62fe !important; /* garde l'anneau bleu */
        background-color: transparent !important;
    }
    div[role="radiogroup"] label input:checked ~ div > div:first-child > div {
        background-color: #0f62fe !important; /* pastille intérieure bleu */
    }
            
    
    /* Buttons */
    div.stButton > button {
        background-color: #0f62fe !important;
        color: white !important;
        border: none !important;
        border-radius: 4px !important;
    }
    div.stButton > button:hover { background-color: #0353e9 !important; }
    div[data-testid="stSidebar"] div.stButton:last-of-type > button {
        background-color: #393939 !important;
        border: 1px solid #525252 !important;
    }

    /* Progress bar color */
    .stProgress > div > div > div > div { background-color: #0f62fe !important; }

    /* badges */
    .badge { padding: 4px 8px; border-radius: 4px; font-weight: 700; font-size: 0.75em; margin-right: 5px; display: inline-block; border: 1px solid transparent; }
    .badge-crit { background-color: transparent !important; color: #da1e28; border-color: transparent !important; }
    .badge-warn { background-color: transparent !important; color: #b8860b; border-color: transparent !important; }
    .badge-safe { background-color: transparent !important; color: #198038; border-color: transparent !important; }
    .badge-info { background-color: transparent !important; color: #0050b3; border-color: transparent !important; }

    /* Header */
    .saas-header {
        background-color: rgb(38, 39, 48); 
        color: white;
        padding: 1.2rem 2rem; 
        border-radius: 8px;
        margin-bottom: 1.5rem;
        display: flex; align-items: center; justify-content: space-between;
        border-top: 4px solid #0f62fe;
        border-bottom: 1px solid rgba(255, 255, 255, 0.1); 
        box-shadow: 0 4px 6px rgba(0,0,0,0.06);
    }

    /* Quote box used for extracted clause snippets */
    .quote-box {
        border-left: 3px solid #0f62fe;
        background-color: rgb(38, 39, 48); 
        padding: 10px;
        font-style: italic;
        color: white;
        margin-top: 5px;
        margin-bottom: 10px;
        font-size: 0.9em;
        border-radius: 4px;
    }

    hr { margin: 8px 0px !important; }
    </style>
    """, unsafe_allow_html=True)

# ==========================================
# 3. BACKEND & watsonx MODEL
# ==========================================
try:
    API_KEY = st.secrets["IBM_API_KEY"]
    PROJECT_ID = st.secrets["IBM_PROJECT_ID"]
except Exception:
    API_KEY = os.getenv("IBM_API_KEY", "")
    PROJECT_ID = os.getenv("IBM_PROJECT_ID", "")

REGION_URL = "https://us-south.ml.cloud.ibm.com"
MODEL_ID = "ibm/granite-3-3-8b-instruct"

@st.cache_resource
def get_model(max_tokens: int = 900):
    """Return a ready ModelInference or None if not configured."""
    if not API_KEY:
        return None
    creds = Credentials(url=REGION_URL, api_key=API_KEY)
    params = {
        GenParams.DECODING_METHOD: "greedy",
        GenParams.MAX_NEW_TOKENS: max_tokens,
        GenParams.MIN_NEW_TOKENS: 5,
        GenParams.STOP_SEQUENCES: ["}"]
    }
    return ModelInference(model_id=MODEL_ID, params=params, credentials=creds, project_id=PROJECT_ID)

def safe_extract_json(text: str):
    """Try multiple strategies to robustly extract the first JSON object from model output."""
    if not isinstance(text, str):
        return None
    try:
        # 1) Triple-backtick JSON block
        match = re.search(r'```json\s*(\{.*?\})\s*```', text, re.DOTALL)
        if match:
            return json.loads(match.group(1))

        # 2) Balanced-brace search (improved to handle nested braces)
        match = re.search(r'\{(?:[^{}]|\{(?:[^{}]|\{[^{}]*\})*\})*\}', text, re.DOTALL)
        if match:
            try:
                return json.loads(match.group(0))
            except json.JSONDecodeError:
                # fallback to incremental decode below
                pass

        # 3) Use JSONDecoder.raw_decode to find a valid JSON substring
        decoder = json.JSONDecoder()
        start = 0
        while True:
            try:
                obj, idx = decoder.raw_decode(text[start:])
                return obj
            except json.JSONDecodeError:
                # move start forward to next brace
                next_brace = text.find("{", start + 1)
                if next_brace == -1:
                    break
                start = next_brace
            except Exception:
                break
    except Exception:
        return None
    return None

def normalize_governing_law(raw: str):
    """Normalize various verbose phrasings into compact jurisdiction names (e.g., 'New York').
    If no clear jurisdiction found, return 'Unknown'."""
    if not raw:
        return "Unknown"
    text = raw.strip()

    # 1) Common explicit patterns
    patterns = [
        r'laws of the State of\s+([A-Za-z\s,]+)',
        r'laws of the state of\s+([A-Za-z\s,]+)',
        r'governed by the laws of the State of\s+([A-Za-z\s,]+)',
        r'governed by the laws of\s+([A-Za-z\s,]+)',
        r'law of the State of\s+([A-Za-z\s,]+)',
        r'law of\s+([A-Za-z\s,]+)',
        r'under the laws of\s+([A-Za-z\s,]+)',
        r'this Agreement shall be governed by the laws of\s+([A-Za-z\s,]+)',
    ]
    for p in patterns:
        m = re.search(p, text, re.IGNORECASE)
        if m:
            candidate = m.group(1).strip().strip(",. ")
            # remove common suffixes and tokens
            candidate = re.sub(r'\b(State|state|Province|province|Commonwealth|the)\b', '', candidate, flags=re.IGNORECASE).strip()
            candidate = re.sub(r'\s+State$', '', candidate, flags=re.IGNORECASE).strip()
            # take first item before any comma
            candidate = candidate.split(",")[0].strip()
            # Title-case
            candidate = " ".join([w.capitalize() for w in candidate.split() if w.strip()])
            if candidate:
                return candidate

    # 2) Look for 'State of X' style
    m2 = re.search(r'(State|state|Province|province|Commonwealth)\s+of\s+([A-Za-z\s]+)', text, re.IGNORECASE)
    if m2:
        candidate = m2.group(2).strip()
        candidate = re.sub(r'\s+State$', '', candidate, flags=re.IGNORECASE).strip()
        return " ".join([w.capitalize() for w in candidate.split()])

    # 3) Look for location after "in" or "located in" patterns e.g. "in New York, USA"
    m3 = re.search(r'in\s+([A-Za-z\s]+?),\s*([A-Za-z\s]{2,})', text)
    if m3:
        # prefer second group if it's a state/country
        candidate = m3.group(2).strip()
        return " ".join([w.capitalize() for w in candidate.split()])

    # 4) As a last attempt, capture the longest capitalized sequence (1-3 words)
    caps = re.findall(r'\b([A-Z][a-z]+(?:\s+[A-Z][a-z]+){0,2})\b', text)
    if caps:
        # choose the longest candidate (by words)
        caps_sorted = sorted(caps, key=lambda s: (-len(s.split()), -len(s)))
        candidate = caps_sorted[0].strip()
        # filter out obvious non-jurisdiction words
        if candidate.lower() not in ("this", "agreement", "party", "parties", "contract", "section"):
            return candidate

    return "Unknown"

@st.cache_data(show_spinner=False)
def analyze_contract(text: str, filename: str):
    """Analyze contract with watsonx model and return structured info. Includes a secondary 'repair' attempt to reduce parsing errors."""
    model = get_model()
    if not model:
        # If model not configured, provide a graceful fallback
        return {
            "risk_score": 0,
            "coc_detected": False,
            "coc_quote": "",
            "assignment_consent_needed": False,
            "assignment_quote": "",
            "governing_law": "API_NOT_CONFIGURED",
            "summary": "Model not configured - set IBM_API_KEY",
            "parties": []
        }

    truncated_text = (text or "")[:15000]  # keep within context
    prompt = f"""
You are an Elite M&A Auditor. Review the contract '{filename}'.

TASK:
### 1. EXTRACTION
- **Parties**: Identify the signing entities.
- **CoC (Change of Control)**: Look for "Termination" rights triggered by a merger/acquisition. Extract QUOTE.
- **Assignment**: Look for "Consent" requirements for transfer. Extract QUOTE.
- **Governing Law**: Extract jurisdiction (Country/State).
### 2. RISK SCORING SCALE (0-100)
Assess the severity of restrictions on the Buyer:
- **80-100 (CRITICAL)**: Explicit "Termination" right upon Change of Control/Merger. (Deal Killer).
- **60-79 (HIGH)**: Assignment is strictly prohibited or "Null and Void" without consent.
- **40-59 (MEDIUM)**: Assignment requires written consent (standard administrative friction).
- **20-39 (LOW)**: Assignment requires only "Notice" to the other party (no consent needed) OR the clause is ambiguous.
- **0-19 (SAFE)**: No restrictions found, or contract explicitly allows assignment.

### OUTPUT (RAW JSON ONLY)
{{
    "risk_score": <int>,
    "coc_detected": <bool>,
    "coc_quote": "<text or empty>",
    "assignment_consent_needed": <bool>,
    "assignment_quote": "<text or empty>",
    "governing_law": "<text or Unknown>",
    "summary": "<Short 1-sentence description in French>",
    "parties": ["Party A", "Party B"]
}}
3. Detect 'Assignment' restrictions (consent needed?). Extract the EXACT QUOTE (if present).
4. Identify the Jurisdiction of Governing Law (e.g., "New York", "France", "Delaware", "England"). 
   - Extract ONLY the State/Country name.
   - If NOT explicitly stated in the text, return "Unknown".
5. Calculate a Risk Score (0-100) for a BUYER. 
   IMPORTANT SCORING RULES:
   - If 'Change of Control' (termination rights) is detected -> Score MUST be between 80 and 100.
   - If only 'Assignment Consent' is needed -> Score between 40 and 60.
   - If no restrictions -> Score between 0 and 10.

OUTPUT RAW JSON ONLY (NO EXPLANATION):
{{
    "risk_score": 0,
    "coc_detected": false,
    "coc_quote": "",
    "assignment_consent_needed": false,
    "assignment_quote": "",
    "governing_law": "",
    "summary": "",
    "parties": []
}}

CONTRACT TEXT:
{truncated_text}
"""
    data = None
    raw_model_output = ""
    try:
        raw_model_output = model.generate_text(prompt)
        data = safe_extract_json(raw_model_output)
    except Exception:
        data = None

    # If parsing failed, attempt a 'repair' pass: ask the model to extract JSON from its previous reply
    if not data:
        try:
            repair_prompt = f"""The previous assistant output failed to parse as JSON. Extract and return ONLY the JSON object present in the text below, matching this schema:

{{"risk_score": <int>,"coc_detected": <bool>,"coc_quote": "<text>","assignment_consent_needed": <bool>,"assignment_quote": "<text>","governing_law": "<text>","summary": "<text>","parties": []}}

TEXT:
{raw_model_output}
"""
            repair_output = model.generate_text(repair_prompt)
            data = safe_extract_json(repair_output)
        except Exception:
            data = None

    if not data:
        # fallback structured response for manual review
        normalized = {
            "risk_score": 50,
            "coc_detected": False,
            "coc_quote": "",
            "assignment_consent_needed": False,
            "assignment_quote": "",
            "governing_law": "Unknown",
            "summary": "Parsing Error - Manual Review Needed",
            "parties": []
        }
        return normalized

    # Ensure fields exist and types are normalized
    try:
        risk_score = int(float(data.get("risk_score", 50))) if data.get("risk_score", None) is not None else 50
    except Exception:
        risk_score = 50

    coc_detected = bool(data.get("coc_detected", False))
    coc_quote = data.get("coc_quote", "") or ""
    assignment_consent_needed = bool(data.get("assignment_consent_needed", False))
    assignment_quote = data.get("assignment_quote", "") or ""
    raw_governing = data.get("governing_law", "") or ""
    governing_law = normalize_governing_law(raw_governing)
    summary_model = (data.get("summary") or "").strip()
    parties = data.get("parties") or []

    # Enforce: if consent is required, ensure at least 30 risk score
    if assignment_consent_needed:
        try:
            if risk_score < 30:
                risk_score = 30
        except Exception:
            risk_score = max(risk_score, 30)

    # --- ALWAYS produce a ONE-SENTENCE summary that describes what the file is ---
    # Build parties string
    if parties and isinstance(parties, (list, tuple)) and len(parties) > 0:
        if len(parties) == 1:
            parties_str = parties[0]
        else:
            # use first two parties for succinctness
            parties_str = " et ".join([p for p in parties[:2]])
    else:
        parties_str = "Parties non spécifiées"

    # Compose a single French sentence describing the file
    sentence_parts = []
    sentence_parts.append(f"Contrat entre {parties_str}")
    if governing_law and governing_law != "Unknown":
        sentence_parts.append(f"régi par {governing_law}")
    # Add risk descriptors
    if coc_detected:
        sentence_parts.append("contient une clause de changement de contrôle")
    elif assignment_consent_needed:
        sentence_parts.append("exige le consentement pour cession")
    else:
        sentence_parts.append("sans restriction de transfert apparente")

    # Join into one sentence
    one_sentence_summary = ", ".join(sentence_parts).strip()
    if not one_sentence_summary.endswith("."):
        one_sentence_summary = one_sentence_summary + "."

    # Use the constructed sentence as summary (ensures consistent one-line description)
    summary = one_sentence_summary[:400]

    normalized = {
        "risk_score": risk_score,
        "coc_detected": coc_detected,
        "coc_quote": coc_quote,
        "assignment_consent_needed": assignment_consent_needed,
        "assignment_quote": assignment_quote,
        "governing_law": governing_law,
        "summary": summary,
        "parties": parties
    }
    return normalized

def generate_letter(filename: str, data: dict, acquiring_entity: str = "[Nom de l'Entité Acquéreuse]"):
    """
    Génère un projet de lettre neutre et professionnel basé sur l'analyse du contrat.
    """
    model = get_model(max_tokens=600)
    if not model:
        return "Erreur : Modèle non configuré."

    # 1. Extraction et Sécurisation du Contexte
    # On récupère les points clés de l'analyse précédente
    assignment_quote = data.get("assignment_quote") or "la clause relative à la cession"
    parties_list = data.get("parties", [])
    parties_str = ", ".join(parties_list) if parties_list else "les parties contractantes"
    
    # 2. Construction du Prompt "Neutre et Contextuel"
    prompt = f"""
    You are an expert legal assistant tasked with drafting professional correspondence.

    CASE CONTEXT:
    - Analyzed Document: '{filename}'
    - Parties Involved: {parties_str}
    - Identified Blocking Clause: "{assignment_quote}"
    - Trigger Event: Merger/Acquisition by {acquiring_entity}

    YOUR TASK:
    Draft a formal letter (email) intended for the counterparty (the other party to the contract).

    DRAFTING GUIDELINES:
    1. **Tone**: Neutral, courteous, direct, and strictly professional. Avoid any emotional or marketing language.
    2. **Objective**: Inform about the merger with {acquiring_entity} and request written consent for the contract assignment, citing the blocking clause as justification.
    3. **Format**: Start directly with "Subject:". Do not add any text before or after the letter body.
    4. **Adaptation**: Use the provided party names to logically identify the recipient.

    DRAFT LETTER:
    """

    try:
        result = model.generate_text(prompt)
        return result.strip()
    except Exception as e:
        return f"Generation Error: {e}"

# ==========================================
# 4. SIDEBAR - UPLOAD & ACTIONS
# ==========================================
with st.sidebar:
    st.divider()

    uploaded_files = st.file_uploader(
        "Upload Contracts (PDF) — drag & drop or click", 
        type=['pdf'], 
        accept_multiple_files=True, 
        key=f"uploader_{st.session_state.uploader_key}"
    )

    if uploaded_files:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("⚡ START AUDIT", type="primary", use_container_width=True):
            bar = st.progress(0)
            status = st.empty()
            success_count = 0
            fail_count = 0

            for i, file in enumerate(uploaded_files):
                status.caption(f"🕵️ Processing: {file.name}")
                try:
                    reader = PyPDF2.PdfReader(file)
                    text = "".join([p.extract_text() or "" for p in reader.pages]) if reader.pages else ""
                    st.session_state.full_corpus += f"\n[[DOC: {file.name}]]\n{text[:3000]}\n"

                    if file.name not in st.session_state.audit_db:
                        time.sleep(0.15)
                        st.session_state.audit_db[file.name] = analyze_contract(text, file.name)
                    success_count += 1
                except Exception:
                    fail_count += 1

                bar.progress((i + 1) / len(uploaded_files))

            st.session_state.audit_stats = {"success": success_count, "fail": fail_count, "total": len(uploaded_files)}
            status.empty()
            bar.empty()
            st.rerun()

    # --- STATS ---
    if st.session_state.audit_stats:
        s = st.session_state.audit_stats
        st.markdown(f"""
        <div style="
            padding:6px; 
            border-radius:5px; 
            border: 1px solid rgba(255, 255, 255, 0.08); 
            margin-bottom:6px; 
            margin-top:6px;
            font-size:0.85rem;
            background-color: transparent; 
            color: white;">
            <div style="margin-bottom:2px; color: white;">✅ <strong>Analyzed:</strong> {s.get('success',0)} files</div>
            <div style="color: white;">❌ <strong>Skipped/Error:</strong> {s.get('fail',0)} files</div>
        </div>
        """, unsafe_allow_html=True)

    # Actions (reset/export) - compact spacing preserved
    # Actions (reset/export)
    if st.session_state.audit_db:
        st.markdown("<hr>", unsafe_allow_html=True)
        st.caption("") # Espaceur
        
        if st.button("🗑️ Delete Cache", use_container_width=True):
            st.session_state.audit_db = {}
            st.session_state.chat_history = []
            st.session_state.full_corpus = ""
            st.session_state.audit_stats = None
            st.session_state.uploader_key += 1
            st.rerun()

        # --- PRÉPARATION DE L'EXPORT PROPRE ---
        # 1. Création du DataFrame brut
        df_export = pd.DataFrame.from_dict(st.session_state.audit_db, orient='index')
        
        # 2. Nettoyage : Sortir l'index (nom du fichier) dans une vraie colonne
        df_export = df_export.reset_index().rename(columns={'index': 'Filename'})
        
        # 3. Nettoyage : Transformer la liste des parties en texte lisible
        if 'parties' in df_export.columns:
            df_export['parties'] = df_export['parties'].apply(lambda x: ", ".join(x) if isinstance(x, list) else str(x))
        
        # 4. Convertir les booléens en OUI/NON pour Excel
        if 'coc_detected' in df_export.columns:
            df_export['coc_detected'] = df_export['coc_detected'].apply(lambda x: "YES" if x else "NO")
        if 'assignment_consent_needed' in df_export.columns:
            df_export['assignment_consent_needed'] = df_export['assignment_consent_needed'].apply(lambda x: "YES" if x else "NO")
        
        # 5. Renommage des colonnes
        rename_map = {
            'risk_score': 'Risk Score (0-100)',
            'coc_detected': 'Change of control ?',
            'coc_quote': 'Clause change of control (Proof)',
            'assignment_consent_needed': 'Consentement cession required ?',
            'assignment_quote': 'Clause cession required (Proof)',
            'governing_law': 'Governing Law',
            'summary': 'Summary',
            'parties': 'Parties',
        }
        df_export = df_export.rename(columns=rename_map)
        
        # 6. Sélection et ordre logique des colonnes
        cols_order = [
            'Filename', 'Risk Score (0-100)', 'Parties', 'Governing Law',
            'Change of control ?', 'Clause change of control (Proof)',
            'Consentement cession required ?', 'Clause cession required (Proof)', 'Summary'
        ]
        final_cols = [c for c in cols_order if c in df_export.columns]
        df_final = df_export[final_cols]
        
        # 7. Export Excel formaté avec openpyxl
        from io import BytesIO
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Créer le fichier Excel en mémoire
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_final.to_excel(writer, index=False, sheet_name='Audit Report')
        
        # Charger le workbook pour le formater
        output.seek(0)
        wb = load_workbook(output)
        ws = wb['Audit Report']
        
        # FORMATAGE DES EN-TÊTES (ligne 1)
        header_fill = PatternFill(start_color="0F62FE", end_color="0F62FE", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # AJUSTER LA LARGEUR DES COLONNES
        column_widths = {
            'A': 30,  # Nom du Fichier
            'B': 20,  # Score Risque
            'C': 25,  # Parties
            'D': 18,  # Droit Applicable
            'E': 25,  # Changement de Contrôle ?
            'F': 50,  # Clause Changement Contrôle (Preuve)
            'G': 28,  # Consentement Cession Requis ?
            'H': 50,  # Clause Cession (Preuve)
            'I': 60   # Résumé Analyse
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # COLORATION CONDITIONNELLE DES SCORES DE RISQUE
        red_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        orange_fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
        green_fill = PatternFill(start_color="E6F7E6", end_color="E6F7E6", fill_type="solid")
        
        for row in range(2, ws.max_row + 1):
            score_cell = ws[f'B{row}']
            try:
                score = int(score_cell.value) if score_cell.value else 0
                if score > 70:
                    score_cell.fill = red_fill  # Rouge pour risque critique
                elif score > 40:
                    score_cell.fill = orange_fill  # Orange pour risque moyen
                else:
                    score_cell.fill = green_fill  # Vert pour risque faible
            except:
                pass
        
        # COLORER LES "OUI" EN ROUGE dans les colonnes CoC et Consentement
        for row in range(2, ws.max_row + 1):
            # Changement de Contrôle ?
            coc_cell = ws[f'E{row}']
            if coc_cell.value == "YES":
                coc_cell.fill = red_fill
                coc_cell.font = Font(bold=True, color="DA1E28")
            
            # Consentement Cession Requis ?
            consent_cell = ws[f'G{row}']
            if consent_cell.value == "YES":
                consent_cell.fill = orange_fill
                consent_cell.font = Font(bold=True, color="B8860B")
        
        # Sauvegarder le fichier formaté
        output_formatted = BytesIO()
        wb.save(output_formatted)
        output_formatted.seek(0)
        
        st.download_button(
            "📥 Excel Export (Audit Report)", 
            output_formatted.getvalue(), 
            "audit_dealflow.xlsx", 
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
            use_container_width=True
        )
# ==========================================
# 5. DASHBOARD / ANALYTICS / GRID / TABS
# ==========================================
# Header
st.markdown("""
<div class="saas-header">
    <div style="display:flex; align-items:center; gap:15px;">
        <img src="https://upload.wikimedia.org/wikipedia/commons/5/51/IBM_logo.svg" width="50">
        <div>
            <div style="font-size:1.4rem; font-weight:800;">DealFlow AI</div>
            <div style="font-size:0.85rem; opacity:0.8;">M&A Due Diligence Suite</div>
        </div>
    </div>
    <div style="text-align:right;">
        <div style="font-size:0.8rem; font-weight:600; color:#0f62fe;">ENTERPRISE EDITION</div>
        <div style="font-size:0.75rem; opacity:0.7;">Powered by <strong>IBM Granite 3.3</strong></div>
    </div>
</div>
""", unsafe_allow_html=True)

# Empty state
if not st.session_state.audit_db:
    st.info("👈 Upload contracts in the sidebar to begin.")
    c1, c2, c3 = st.columns(3)
    c1.markdown("##### 1. Upload"); c1.caption("Drag & Drop PDFs 📁")
    c2.markdown("##### 2. AI Audit"); c2.caption("Detect Deal Breakers 🛡️")
    c3.markdown("##### 3. Action"); c3.caption("Generate Legal Docs 📧")
else:
    # Build dataframe safely
    df = pd.DataFrame.from_dict(st.session_state.audit_db, orient='index').reset_index().rename(columns={'index': 'Document'})
    # Ensure expected columns exist (keep backend field assignment_consent_needed)
    expected_cols = ["risk_score", "coc_detected", "assignment_consent_needed", "governing_law", "summary", "coc_quote", "assignment_quote", "parties"]
    for col in expected_cols:
        if col not in df.columns:
            df[col] = None

    # Convert types for charts/metrics
    df["risk_score"] = pd.to_numeric(df["risk_score"].fillna(0), errors='coerce').fillna(0).astype(int)
    df["coc_detected"] = df["coc_detected"].fillna(False).astype(bool)
    df["assignment_consent_needed"] = df["assignment_consent_needed"].fillna(False).astype(bool)
    df["governing_law"] = df["governing_law"].fillna("Unknown")
    df["summary"] = df["summary"].fillna("")

    # --- 1) Sort PDFs by danger (risk_score desc) for the dashboard ---
    df = df.sort_values(by="risk_score", ascending=False).reset_index(drop=True)

    # Top KPIs
    k1, k2, k3, k4 = st.columns(4)
    high_risk = len(df[df["risk_score"] > 70])
    coc_count = int(df["coc_detected"].sum())
    avg_risk = int(df["risk_score"].mean()) if not df["risk_score"].empty else 0

    k1.metric("Documents", len(df))
    k2.metric("Deal Breakers", high_risk, delta="Requires Attention" if high_risk > 0 else "Safe")
    k3.metric("Change of Control", coc_count, delta="Deal Breakers")
    k4.metric("Avg Risk", f"{avg_risk}/100")

    st.divider()

    # Portfolio analytics (charts)
    st.subheader("📊 Portfolio Analytics")
    c1, c2 = st.columns([1, 1])
    with c1:
        chart_risk = alt.Chart(df).mark_bar().encode(
            x=alt.X("risk_score:Q", bin=alt.Bin(maxbins=10), title="Risk Score"),
            y=alt.Y("count()", title="Count"),
            color=alt.Color('risk_score:Q', scale=alt.Scale(scheme='redyellowgreen'), legend=None),
            tooltip=[alt.Tooltip("count()", title="Count")]
        ).properties(title="Risk Distribution", height=220)
        st.altair_chart(chart_risk, use_container_width=True)
    with c2:
        # Group by governing law - small handling for many categories
        law_counts = df.groupby("governing_law").size().reset_index(name='count')
        if not law_counts.empty:
            chart_law = alt.Chart(law_counts).mark_arc(innerRadius=40).encode(
                theta=alt.Theta("count:Q", title="Count"),
                color=alt.Color("governing_law:N", title="Governing Law"),
                tooltip=["governing_law", "count"]
            ).properties(title="Governing Law Jurisdiction", height=220)
            st.altair_chart(chart_law, use_container_width=True)
        else:
            st.write("No governing law data available.")

    st.subheader("🗃️ Smart Audit Matrix")
    # Build a view for st.data_editor if available
    try:
        # Re-added the "Consent Req." column to the visible grid as requested
        column_cfg = {
            "risk_score": st.column_config.ProgressColumn("Risk Score", format="%d", min_value=0, max_value=100, width="small"),
            "coc_detected": st.column_config.CheckboxColumn("CoC Risk", width="small"),
            "assignment_consent_needed": st.column_config.CheckboxColumn("Consent Req.", width="small"),
            "summary": st.column_config.TextColumn("AI Summary", width="large"),
            "governing_law": st.column_config.TextColumn("Governing Law", width="small")
        }
        # Build display_df including the assignment_consent_needed column
        display_df = df[["Document", "risk_score", "coc_detected", "assignment_consent_needed", "governing_law", "summary"]]
        _ = st.data_editor(display_df, column_config=column_cfg, use_container_width=True, hide_index=True, height=300)
    except Exception:
        # Fallback simple table (including the consent column)
        st.dataframe(df[["Document", "risk_score", "coc_detected", "assignment_consent_needed", "governing_law", "summary"]], use_container_width=True)

    st.markdown("---")
    tab1, tab2 = st.tabs(["🔍 Deep Dive & Evidence", "💬 Legal Assistant (RAG)"])

    with tab1:
        st.info("Select a document to verify AI findings against the original text.")
        # --- Filters as requested: ALL / Consent Required / Critical files ---
        filter_opt = st.radio("", ["ALL", "Consent Required", "Critical files"], horizontal=True)
        # Iterate in sorted order (already sorted above)
        for _, row in df.iterrows():
            filename = row["Document"]
            data = st.session_state.audit_db.get(filename, {})
            # --- Apply filter logic ---
            if filter_opt == "Consent Required" and not data.get("assignment_consent_needed"):
                continue
            if filter_opt == "Critical files" and not (int(row["risk_score"]) > 70):
                continue

            score = int(row["risk_score"])
            is_critical = score > 70
            # Use a red pastille icon for critical files
            icon = "🔴" if is_critical else "🟢"
            with st.expander(f"{icon} {filename} — Risk: {score}", expanded=is_critical):
                c_info, c_evidence = st.columns([1, 2])
                with c_info:
                    st.markdown(f"**Governing Law:** `{data.get('governing_law', 'Unknown')}`")
                    parties = data.get("parties") or []
                    st.markdown(f"**Parties:** {', '.join(parties) if parties else 'N/A'}")
                    st.caption(data.get("summary", ""))
                    # Keep existing functionality: draft button still appears when backend flag is true
                    if data.get("assignment_consent_needed"):
                        if st.button(f"✉️ Draft Consent Letter", key=f"btn_{filename}"):
                            with st.spinner("Agent writing..."):
                                st.session_state[f"let_{filename}"] = generate_letter(filename, data)
                with c_evidence:
                    if data.get("coc_detected"):
                        st.markdown("**⚠️ Change of Control Clause Detected:**")
                        st.markdown(f"<div class='quote-box'>“{data.get('coc_quote','No text extracted')}”</div>", unsafe_allow_html=True)
                    if data.get("assignment_consent_needed"):
                        st.markdown("**✋ Assignment Consent Clause:**")
                        st.markdown(f"<div class='quote-box'>“{data.get('assignment_quote','No text extracted')}”</div>", unsafe_allow_html=True)
                    if f"let_{filename}" in st.session_state:
                        # Show draft
                        draft_text = st.session_state[f"let_{filename}"]
                        st.text_area("Generated Draft:", draft_text, height=200)
                        # --- Copy button under draft (no alert) ---
                        escaped = (draft_text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                                   .replace('"', "&quot;").replace("'", "&#39;"))
                        safe_id = re.sub(r'\W+', '_', filename)
                        html = f"""
                        <div style="margin-top:8px;">
                          <textarea id="hidden_{safe_id}" style="display:none;">{escaped}</textarea>
                          <button id="copybtn_{safe_id}">Copy Draft</button>
                          <span id="copied_{safe_id}" style="display:none; margin-left:10px; color:green;">Copied</span>
                          <span id="err_{safe_id}" style="display:none; margin-left:10px; color:red;">Copy failed</span>
                        </div>
                        <script>
                        const btn = document.getElementById('copybtn_{safe_id}');
                        const hidden = document.getElementById('hidden_{safe_id}');
                        const ok = document.getElementById('copied_{safe_id}');
                        const err = document.getElementById('err_{safe_id}');

                        btn.addEventListener('click', async () => {{
                          ok.style.display = 'none';
                          err.style.display = 'none';
                          try {{
                            if (navigator.clipboard && navigator.clipboard.writeText) {{
                              await navigator.clipboard.writeText(hidden.value);
                            }} else {{
                              // fallback for older browsers
                              hidden.style.display = 'block';
                              hidden.select();
                              document.execCommand('copy');
                              hidden.style.display = 'none';
                            }}
                            ok.style.display = 'inline-block';
                            // auto-hide after 2s
                            setTimeout(() => {{ ok.style.display = 'none'; }}, 2000);
                          }} catch (e) {{
                            err.style.display = 'inline-block';
                            setTimeout(() => {{ err.style.display = 'none'; }}, 3000);
                          }}
                        }});
                        </script>
                        """
                        components.html(html, height=90)

    with tab2:
        st.markdown("#### 🧠 Ask across uploaded documents")
        st.caption("RAG-style quick answers using metadata. (Full RAG with embeddings not included.)")
        for msg in st.session_state.chat_history:
            st.chat_message(msg["role"]).write(msg["content"])

        if prompt := st.chat_input("Ex: Which contracts are under New York law?"):
            st.session_state.chat_history.append({"role": "user", "content": prompt})
            st.chat_message("user").write(prompt)
            with st.chat_message("assistant"):
                with st.spinner("Thinking..."):
                    # Build a lightweight context from metadata
                    context_lines = []
                    for fn, dd in st.session_state.audit_db.items():
                        context_lines.append(f"DOC: {fn} | LAW: {dd.get('governing_law','Unknown')} | RISK: {dd.get('risk_score',0)} | SUM: {dd.get('summary','')}")
                    context = "\n".join(context_lines)
                    model = get_model(max_tokens=600)
                    if not model:
                        res = "Model not configured. Set IBM_API_KEY to enable assistant."
                    else:
                        full_prompt = f"Context of contracts:\n{context}\n\nUser Question: {prompt}\nAnswer concisely as an M&A lawyer, referencing documents if relevant."
                        try:
                            res = model.generate_text(full_prompt)
                        except Exception as e:
                            res = f"Error from model: {e}"
                    st.write(res)
                    st.session_history = st.session_state  # no-op line to avoid lint issues; state is preserved
                    st.session_state.chat_history.append({"role": "assistant", "content": res})